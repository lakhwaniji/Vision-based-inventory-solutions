import openpyxl


def createfile(filename):
    workbook = openpyxl.Workbook()
    work = workbook.active
    work.column_dimensions['A'].width = 20
    work.column_dimensions['B'].width = 20
    work.column_dimensions['C'].width = 20
    workbook.save(filename=filename)


def writedata(filename, name, regno, sid):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        data = (name, regno, sid)
        sheet.append(data)
        wb.save(filename)
        return "Success"
    except:
        return "Error"


def readcardno(filename):
    try:
        wb2 = openpyxl.load_workbook("IOT_SEC_A.xlsx")
        sheet = wb2.active
        values = [sheet.cell(row=i, column=3).value for i in range(1, sheet.max_row + 1)]
        return values
    except:
        return "Error"


def readstudname(filename):
    try:
        wb2 = openpyxl.load_workbook("IOT_SEC_A.xlsx")
        sheet = wb2.active
        values = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
        return values
    except:
        return "Error"


def readregno(filename):
    try:
        wb2 = openpyxl.load_workbook("IOT_SEC_A.xlsx")
        sheet = wb2.active
        values = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row + 1)]
        return values
    except:
        return "Error"


def writecomponent(filename, regno, comp_id):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        data = (regno, comp_id)
        sheet.append(data)
        wb.save(filename)
        return "Success"
    except:
        return "Error"


def getcomponents(card_number):
    wb2 = openpyxl.load_workbook("stock_data.xlsx")
    sheet = wb2.active
    components = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row + 1)]
    card_data = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
    result = ""
    for i in range(0, len(card_data)):
        if (card_data[i] == card_number):
            result = result + components[i] + "  "
    return result


def deletedata(card_number):
    wb2 = openpyxl.load_workbook("stock_data.xlsx")
    sheet = wb2.active
    for idxs, rows in enumerate(sheet['A']):
        if (rows.value == card_number):
            wb2.save("stock_data.xlsx")
            return idxs + 1
    wb2.save("stock_data.xlsx")
    return False


def delete(card_number):
    idx = deletedata(card_number)
    if (idx == False):
        return "Success"
    else:
        wb2 = openpyxl.load_workbook("stock_data.xlsx")
        sheet = wb2.active
        sheet.delete_rows(idx=idx, amount=1)
        wb2.save("stock_data.xlsx")
        delete(card_number)
    return "Success"
